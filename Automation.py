import os
import pdfplumber
import re
import win32com.client
import openpyxl
import datetime
import csv
from copy import copy

current_folder = os.path.dirname(os.path.abspath(__file__))

# --- MATHEMATICAL SAFEGUARDS & HELPERS ---
def is_cell_merged(ws, row, col):
    """Safely checks if a specific coordinate is part of any merged cell range."""
    for merged_range in ws.merged_cells.ranges:
        if row >= merged_range.min_row and row <= merged_range.max_row and col >= merged_range.min_col and col <= merged_range.max_col:
            return True
    return False

def get_true_bottom_row(ws, min_row=10):
    """Finds the absolute last row that contains a ticket number or name."""
    last_row = min_row - 1
    for r in range(min_row, ws.max_row + 1):
        if ws.cell(row=r, column=3).value is not None or ws.cell(row=r, column=4).value is not None:
            last_row = r
    return last_row

def get_next_safe_row(ws, start_row):
    """Finds the next row that has no data in Col 3/4 and contains NO merged cells from Cols 3 to 10."""
    current_row = start_row
    while True:
        if ws.cell(row=current_row, column=3).value is not None or ws.cell(row=current_row, column=4).value is not None:
            current_row += 1
            continue
            
        row_has_merge = False
        for c in range(3, 11):
            if is_cell_merged(ws, current_row, c):
                row_has_merge = True
                break
                
        if row_has_merge:
            current_row += 1
        else:
            return current_row

def get_month(val):
    """Extracts the numerical month from various date formats."""
    if not val: 
        return None
    if isinstance(val, datetime.datetime):
        return val.month
    
    val_str = str(val).strip()
    match = re.search(r'^(\d{1,2})[/-]\d{1,2}[/-]\d{2,4}', val_str)
    if match: return int(match.group(1))
    match = re.search(r'^\d{4}[/-](\d{1,2})[/-]\d{1,2}', val_str)
    if match: return int(match.group(1))
    
    try:
        dt = datetime.datetime.strptime(val_str.replace('"', ''), "%b %d, %Y")
        return dt.month
    except ValueError:
        pass

    return None

def parse_csv_date(date_str):
    """Parses date string for sorting purposes."""
    if not date_str:
        return datetime.datetime.min
    try:
        return datetime.datetime.strptime(date_str.strip().replace('"', ''), "%b %d, %Y")
    except ValueError:
        return datetime.datetime.min

def parse_date_to_obj(date_str):
    """Converts CSV date strings to true datetime objects for native Excel formatting."""
    if not date_str:
        return None
    try:
        return datetime.datetime.strptime(date_str.strip().replace('"', ''), "%b %d, %Y")
    except ValueError:
        return None

def apply_row_formatting(ws, template_row, target_row, start_col, end_col):
    """Copies font, border, fill, alignment, and number formats from a template row to blend in with existing data."""
    for col in range(start_col, end_col + 1):
        source_cell = ws.cell(row=template_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        if source_cell.has_style:
            if source_cell.font: target_cell.font = copy(source_cell.font)
            if source_cell.border: target_cell.border = copy(source_cell.border)
            if source_cell.fill: target_cell.fill = copy(source_cell.fill)
            if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
            if source_cell.number_format: target_cell.number_format = source_cell.number_format
# --------------------------------

def run_valet_automation():
    print(f"--- STARTING AUTOMATION: {datetime.datetime.now()} ---")
    downloaded_files = []
    
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) 
        root_folder = inbox.Parent
        
        target_folder = root_folder.Folders.Item("Revenue Tracking")
        print(f"TROUBLESHOOT: Accessed folder: {target_folder.Name}")

        now = datetime.datetime.now()
        first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = first_of_month.strftime("%m/%d/%Y %I:%M %p")
        print(f"TROUBLESHOOT: Searching for emails received on or after {date_filter}")
        
        sender_email = "noreply@notifications.parkingmgt.com"
        sFilter = f"[SenderEmailAddress] = '{sender_email}' AND [ReceivedTime] >= '{date_filter}'"
        
        messages = target_folder.Items.Restrict(sFilter)
        messages.Sort("[ReceivedTime]", True)
        
        print(f"TROUBLESHOOT: Found {messages.Count} potential emails in Revenue Tracking.")

        found_recent_step2 = False
        found_recent_step3 = False

        for msg in messages:
            try:
                subject = msg.Subject
                subject_lower = subject.lower()
                
                if "trilogy" in subject_lower and "huntsville" in subject_lower:
                    if "step#1" in subject_lower or "step 1" in subject_lower:
                        prefix = "Step1_"
                    elif "step#2" in subject_lower or "step 2" in subject_lower:
                        if found_recent_step2:
                            continue  
                        prefix = "Step2_"
                        found_recent_step2 = True
                    elif "validation summary report" in subject_lower:
                        if found_recent_step3:
                            continue
                        prefix = "Step3_"
                        found_recent_step3 = True
                    else:
                        continue 

                    received_date = msg.ReceivedTime.strftime("%Y-%m-%d")
                    has_attachment = False
                    
                    for attachment in msg.Attachments:
                        if attachment.FileName.lower().endswith((".pdf", ".csv")):
                            has_attachment = True
                            unique_filename = f"{prefix}{received_date}_{attachment.FileName}"
                            pdf_path = os.path.join(current_folder, unique_filename)
                            
                            if not os.path.exists(pdf_path):
                                attachment.SaveAsFile(pdf_path)
                                downloaded_files.append(pdf_path)
                                print(f"SUCCESS: Downloaded '{unique_filename}'")
                            else:
                                downloaded_files.append(pdf_path)
                            break 
                            
            except Exception as e:
                pass 
        
        # --- D184 DOWNLOAD LOGIC ---
        print(f"TROUBLESHOOT: Searching Inbox for D184 emails...")
        inbox_messages = inbox.Items
        inbox_messages.Sort("[ReceivedTime]", True)
        
        d184_pdf_path = None
        for msg in inbox_messages:
            try:
                subject_lower = msg.Subject.lower() if msg.Subject else ""
                
                for attachment in msg.Attachments:
                    filename_lower = attachment.FileName.lower()
                    if filename_lower.endswith(".pdf") and ("d184" in subject_lower or "d184" in filename_lower):
                        current_date_str = datetime.datetime.now().strftime("%Y-%m-%d")
                        unique_filename = f"D184-{current_date_str}.pdf"
                        full_path = os.path.join(current_folder, unique_filename)
                        
                        if not os.path.exists(full_path):
                            attachment.SaveAsFile(full_path)
                            print(f"SUCCESS: Downloaded '{unique_filename}' from Inbox")
                        else:
                            print(f"TROUBLESHOOT: '{unique_filename}' already exists. Using existing file.")
                        
                        d184_pdf_path = full_path
                        break 
                
                if d184_pdf_path:
                    downloaded_files.append(d184_pdf_path)
                    break 
            except Exception as e:
                pass

    except Exception as e:
        print(f"TROUBLESHOOT ERROR (Outlook): {e}")

    if not downloaded_files:
        print("RESULT: No new data found in Outlook. Exiting.")
        return

    excel_path = r"C:\Users\AltonJones\Desktop\Automation test\2026 RSS-102742 Trilogy Hotel Huntsville.xlsx"
    
    try:
        print(f"TROUBLESHOOT: Opening workbook at {excel_path}")
        wb_save = openpyxl.load_workbook(excel_path)

        for filepath in downloaded_files:
            filename = os.path.basename(filepath)
            print(f"\n--- Processing: {filename} ---")
            
            # --- STEP #1 LOGIC ---
            if filename.startswith("Step1_") and filename.endswith(".pdf"):
                with pdfplumber.open(filepath) as pdf:
                    text = pdf.pages[0].extract_text()
                    
                date_match = re.search(r'Overnight Revenue Summary\s+([A-Z][a-z]{2})\s+(\d{1,2})', text, re.IGNORECASE)
                if not date_match: 
                    continue

                month_str = date_match.group(1).capitalize()
                day_num = int(date_match.group(2))
                
                vehicles = re.findall(r'Vehicle Count\s+(\d+)', text)
                revenues = re.findall(r'Revenue\s+\$?([\d,]+\.\d{2})', text)
                
                if len(vehicles) >= 2 and len(revenues) >= 2:
                    dv = int(vehicles[0])
                    dr = float(revenues[0].replace(',', ''))
                    ov = int(vehicles[1])
                    or_rev = float(revenues[1].replace(',', ''))

                    sheet_map = {"Jan": "RSS_Jan", "Feb": "RSS_Feb", "Mar": "RSS_Mar", "Apr": "RSS_Apr", "May": "RSS_May", "Jun": "RSS_June", "Jul": "RSS_July", "Aug": "RSS_Aug", "Sep": "RSS_Sept", "Oct": "RSS_Oct", "Nov": "RSS_Nov", "Dec": "RSS_Dec"}
                    sheet_name = sheet_map.get(month_str, f"RSS_{month_str}")
                    
                    if sheet_name in wb_save.sheetnames:
                        ws = wb_save[sheet_name]
                        
                        dv_row, dr_row = 73, 74
                        ov_row, or_rev_row = 79, 80
                        
                        for row in range(1, 150):
                            val_c = str(ws.cell(row=row, column=3).value).strip()
                            if val_c == "Valet Daily Revenue Collected by PMC":
                                dv_row, dr_row = row, row + 1
                            elif val_c == "Valet Overnight Revenue Collected by PMC":
                                ov_row, or_rev_row = row, row + 1

                        target_col = day_num + 4
                        
                        if not is_cell_merged(ws, dv_row, target_col): ws.cell(row=dv_row, column=target_col).value = dv
                        if not is_cell_merged(ws, dr_row, target_col): ws.cell(row=dr_row, column=target_col).value = dr
                        if not is_cell_merged(ws, ov_row, target_col): ws.cell(row=ov_row, column=target_col).value = ov
                        if not is_cell_merged(ws, or_rev_row, target_col): ws.cell(row=or_rev_row, column=target_col).value = or_rev
                            
                        print(f"SUCCESS: Daily & Overnight Data for {month_str} {day_num} updated in Column {target_col}")

            # --- D184 LOGIC ---
            elif filename.startswith("D184-") and filename.endswith(".pdf"):
                sheet_map_by_num = {1: "RSS_Jan", 2: "RSS_Feb", 3: "RSS_Mar", 4: "RSS_Apr", 5: "RSS_May", 6: "RSS_June", 7: "RSS_July", 8: "RSS_Aug", 9: "RSS_Sept", 10: "RSS_Oct", 11: "RSS_Nov", 12: "RSS_Dec"}
                daily_totals = {}
                
                with pdfplumber.open(filepath) as pdf:
                    for page in pdf.pages:
                        table = page.extract_table()
                        if table:
                            for row in table:
                                if row and len(row) >= 5 and row[0]:
                                    date_str = str(row[0]).strip().replace('\n', '')
                                    match = re.match(r'^(\d{2})-(\d{2})-(\d{2})$', date_str)
                                    if match:
                                        month_num = int(match.group(1))
                                        day_num = int(match.group(2))
                                        amount_str = str(row[-1]).strip().replace('$', '').replace(',', '')
                                        try:
                                            amount = float(amount_str)
                                            key = (month_num, day_num)
                                            daily_totals[key] = daily_totals.get(key, 0.0) + amount
                                        except ValueError:
                                            pass
                        else:
                            # Fallback if table extraction fails
                            text = page.extract_text()
                            if text:
                                for line in text.split('\n'):
                                    match = re.search(r'^(\d{2})-(\d{2})-(\d{2})\s+.*\s+([\d,]+\.\d{2})$', line.strip())
                                    if match:
                                        month_num = int(match.group(1))
                                        day_num = int(match.group(2))
                                        amount = float(match.group(4).replace(',', ''))
                                        key = (month_num, day_num)
                                        daily_totals[key] = daily_totals.get(key, 0.0) + amount

                for (month_num, day_num), total_amount in daily_totals.items():
                    sheet_name = sheet_map_by_num.get(month_num)
                    if sheet_name and sheet_name in wb_save.sheetnames:
                        ws = wb_save[sheet_name]
                        target_col = day_num + 4
                        target_row = 36
                        
                        if not is_cell_merged(ws, target_row, target_col):
                            ws.cell(row=target_row, column=target_col).value = total_amount
                            print(f"SUCCESS: Wrote D184 total ${total_amount:.2f} for {sheet_name} Day {day_num} to Row {target_row}, Column {target_col}")

            # --- STEP #2 LOGIC ---
            elif filename.startswith("Step2_"):
                sheet_name = "Session Discrepancy"
                if sheet_name in wb_save.sheetnames:
                    ws_disc = wb_save[sheet_name]
                    
                    existing_tickets = set()
                    for r in range(10, ws_disc.max_row + 1):
                        val = ws_disc.cell(row=r, column=3).value
                        if val is not None:
                            if isinstance(val, float): existing_tickets.add(str(int(val)))
                            else: existing_tickets.add(str(val).strip())

                    last_data_row = get_true_bottom_row(ws_disc, 10)
                    last_month = None
                    if last_data_row >= 10:
                        last_arrival_val = ws_disc.cell(row=last_data_row, column=5).value
                        last_month = get_month(last_arrival_val)

                    next_empty_row = get_next_safe_row(ws_disc, last_data_row + 1)
                    extracted_data = []
                    
                    if filename.endswith(".pdf"):
                        with pdfplumber.open(filepath) as pdf:
                            for page in pdf.pages:
                                table = page.extract_table()
                                if table: extracted_data.extend(table)
                    else:
                        with open(filepath, newline='', encoding='utf-8') as f:
                            reader = csv.reader(f)
                            extracted_data = list(reader)

                    headers = []
                    processing_overnights = False
                    
                    for row_data in extracted_data:
                        row_data = [str(x).strip().replace('\n', ' ') if x else "" for x in row_data]
                        
                        if all(x == "" for x in row_data): continue
                            
                        row_text = " ".join(row_data)
                        if "Overnight Valet" in row_text:
                            processing_overnights = True
                            continue
                        elif "Daily Valet" in row_text:
                            break
                            
                        if not processing_overnights: continue
                        if "Ticket" in row_data and "Guest Name" in row_data:
                            headers = row_data
                            continue
                        if "Total" in row_data: continue
                            
                        if headers and len(row_data) > 0 and row_data[0].isdigit():
                            ticket = row_data[headers.index("Ticket")] if "Ticket" in headers else ""
                            
                            ticket_str = str(ticket).strip()
                            if ticket_str in existing_tickets: continue
                            
                            guest_name = row_data[headers.index("Guest Name")] if "Guest Name" in headers else ""
                            arrival = row_data[headers.index("Arrival")] if "Arrival" in headers else ""
                            
                            depart = ""
                            if "Departure" in headers: depart = row_data[headers.index("Departure")]
                            elif "Depart" in headers: depart = row_data[headers.index("Depart")]
                            
                            room = ""
                            if "Room Number" in headers and headers.index("Room Number") < len(row_data): 
                                room = row_data[headers.index("Room Number")]
                                
                            try:
                                room_val = int(room)
                            except ValueError:
                                try:
                                    room_val = float(room)
                                except ValueError:
                                    room_val = room
                            
                            collected = "0"
                            if "Collected" in headers and headers.index("Collected") < len(row_data): 
                                collected = row_data[headers.index("Collected")].replace("$","").replace(",","")
                            
                            uncollected = "0"
                            if "Uncollected" in headers and headers.index("Uncollected") < len(row_data): 
                                uncollected = row_data[headers.index("Uncollected")].replace("$","").replace(",","")
                            
                            current_month = get_month(arrival)
                            if current_month and last_month and current_month != last_month:
                                next_empty_row = get_next_safe_row(ws_disc, next_empty_row + 1)
                            
                            if current_month: last_month = current_month

                            ws_disc.cell(row=next_empty_row, column=3).value = int(ticket) if ticket.isdigit() else ticket
                            ws_disc.cell(row=next_empty_row, column=4).value = guest_name
                            ws_disc.cell(row=next_empty_row, column=5).value = arrival
                            ws_disc.cell(row=next_empty_row, column=6).value = depart
                            ws_disc.cell(row=next_empty_row, column=7).value = room_val
                            
                            try: ws_disc.cell(row=next_empty_row, column=8).value = float(collected)
                            except ValueError: pass
                                
                            try: ws_disc.cell(row=next_empty_row, column=9).value = float(uncollected)
                            except ValueError: pass

                            apply_row_formatting(ws_disc, 10, next_empty_row, 3, 9)
                            
                            print(f"SUCCESS: Wrote ticket {ticket} to safe row {next_empty_row}")
                            
                            existing_tickets.add(ticket_str)
                            next_empty_row = get_next_safe_row(ws_disc, next_empty_row + 1)

            # --- STEP #3 LOGIC (BTR Tracker) ---
            elif filename.startswith("Step3_") and filename.endswith(".csv"):
                sheet_name = "BTR Tracker"
                if sheet_name in wb_save.sheetnames:
                    ws_btr = wb_save[sheet_name]
                    
                    existing_names = set()
                    for r in range(10, ws_btr.max_row + 1):
                        val = ws_btr.cell(row=r, column=4).value
                        if val is not None:
                            existing_names.add(str(val).strip().lower())

                    last_data_row = get_true_bottom_row(ws_btr, 10)
                    last_month = None
                    if last_data_row >= 10:
                        last_arrival_val = ws_btr.cell(row=last_data_row, column=6).value
                        last_month = get_month(last_arrival_val)

                    next_empty_row = get_next_safe_row(ws_btr, last_data_row + 1)

                    extracted_data = []
                    with open(filepath, newline='', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            extracted_data.append(row)

                    valid_rows = []
                    for row in extracted_data:
                        if len(row) >= 9 and row[1].strip().isdigit():
                            validation_str = row[6].strip().replace('\n', ' ')
                            if "bill to room 2026 (btr26)" in validation_str.lower():
                                valid_rows.append(row)

                    valid_rows.sort(key=lambda x: (
                        x[6].strip(), 
                        parse_csv_date(x[4]), 
                        parse_csv_date(x[5]), 
                        int(x[1].strip())
                    ))

                    for row_data in valid_rows:
                        guest_name = row_data[2].strip()
                        name_key = guest_name.lower()
                        
                        if name_key in existing_names:
                            print(f"TROUBLESHOOT: Skipping duplicate name '{guest_name}'")
                            continue

                        ticket_str = row_data[1].strip()
                        room = row_data[3].strip()
                        
                        try:
                            room_val = int(room)
                        except ValueError:
                            try:
                                room_val = float(room)
                            except ValueError:
                                room_val = room
                                
                        arrival_raw = row_data[4].strip()
                        depart_raw = row_data[5].strip()
                        validation = row_data[6].strip().replace('\n', ' ')
                        amount = row_data[8].strip().replace('$', '').replace(',', '')
                        
                        arrival_dt = parse_date_to_obj(arrival_raw)
                        depart_dt = parse_date_to_obj(depart_raw)

                        current_month = get_month(arrival_raw)
                        if current_month and last_month and current_month != last_month:
                            print(f"TROUBLESHOOT: Month boundary detected in BTR ({last_month} -> {current_month}). Adding blank row.")
                            next_empty_row = get_next_safe_row(ws_btr, next_empty_row + 1)
                        
                        if current_month:
                            last_month = current_month

                        ws_btr.cell(row=next_empty_row, column=3).value = int(ticket_str)         
                        ws_btr.cell(row=next_empty_row, column=4).value = guest_name              
                        ws_btr.cell(row=next_empty_row, column=5).value = room_val                
                        ws_btr.cell(row=next_empty_row, column=6).value = arrival_dt if arrival_dt else arrival_raw 
                        ws_btr.cell(row=next_empty_row, column=7).value = depart_dt if depart_dt else depart_raw   
                        ws_btr.cell(row=next_empty_row, column=8).value = validation              
                        ws_btr.cell(row=next_empty_row, column=9).value = "Yes"                   
                        
                        try:
                            ws_btr.cell(row=next_empty_row, column=10).value = float(amount)      
                        except ValueError:
                            ws_btr.cell(row=next_empty_row, column=10).value = amount
                        
                        apply_row_formatting(ws_btr, 10, next_empty_row, 3, 10)

                        print(f"SUCCESS: Wrote BTR ticket for {guest_name} to safe row {next_empty_row}")
                        
                        existing_names.add(name_key) 
                        next_empty_row = get_next_safe_row(ws_btr, next_empty_row + 1)

        wb_save.save(excel_path)
        wb_save.close()
        
        print("\nSUCCESS: Workbook saved successfully. Cleaning up temp files...")

        for filepath in downloaded_files:
            try: 
                os.remove(filepath)
            except: pass
            
        print("Done.")

    except Exception as e:
        import traceback
        print(f"CRITICAL ERROR: {e}")
        traceback.print_exc() 

if __name__ == "__main__":
    run_valet_automation()import os
import pdfplumber
import re
import win32com.client
import openpyxl
import datetime
import csv
from copy import copy

current_folder = os.path.dirname(os.path.abspath(__file__))

# --- MATHEMATICAL SAFEGUARDS & HELPERS ---
def is_cell_merged(ws, row, col):
    """Safely checks if a specific coordinate is part of any merged cell range."""
    for merged_range in ws.merged_cells.ranges:
        if row >= merged_range.min_row and row <= merged_range.max_row and col >= merged_range.min_col and col <= merged_range.max_col:
            return True
    return False

def get_true_bottom_row(ws, min_row=10):
    """Finds the absolute last row that contains a ticket number or name."""
    last_row = min_row - 1
    for r in range(min_row, ws.max_row + 1):
        if ws.cell(row=r, column=3).value is not None or ws.cell(row=r, column=4).value is not None:
            last_row = r
    return last_row

def get_next_safe_row(ws, start_row):
    """Finds the next row that has no data in Col 3/4 and contains NO merged cells from Cols 3 to 10."""
    current_row = start_row
    while True:
        if ws.cell(row=current_row, column=3).value is not None or ws.cell(row=current_row, column=4).value is not None:
            current_row += 1
            continue
            
        row_has_merge = False
        for c in range(3, 11):
            if is_cell_merged(ws, current_row, c):
                row_has_merge = True
                break
                
        if row_has_merge:
            current_row += 1
        else:
            return current_row

def get_month(val):
    """Extracts the numerical month from various date formats."""
    if not val: 
        return None
    if isinstance(val, datetime.datetime):
        return val.month
    
    val_str = str(val).strip()
    match = re.search(r'^(\d{1,2})[/-]\d{1,2}[/-]\d{2,4}', val_str)
    if match: return int(match.group(1))
    match = re.search(r'^\d{4}[/-](\d{1,2})[/-]\d{1,2}', val_str)
    if match: return int(match.group(1))
    
    try:
        dt = datetime.datetime.strptime(val_str.replace('"', ''), "%b %d, %Y")
        return dt.month
    except ValueError:
        pass

    return None

def parse_csv_date(date_str):
    """Parses date string for sorting purposes."""
    if not date_str:
        return datetime.datetime.min
    try:
        return datetime.datetime.strptime(date_str.strip().replace('"', ''), "%b %d, %Y")
    except ValueError:
        return datetime.datetime.min

def parse_date_to_obj(date_str):
    """Converts CSV date strings to true datetime objects for native Excel formatting."""
    if not date_str:
        return None
    try:
        return datetime.datetime.strptime(date_str.strip().replace('"', ''), "%b %d, %Y")
    except ValueError:
        return None

def apply_row_formatting(ws, template_row, target_row, start_col, end_col):
    """Copies font, border, fill, alignment, and number formats from a template row to blend in with existing data."""
    for col in range(start_col, end_col + 1):
        source_cell = ws.cell(row=template_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        if source_cell.has_style:
            if source_cell.font: target_cell.font = copy(source_cell.font)
            if source_cell.border: target_cell.border = copy(source_cell.border)
            if source_cell.fill: target_cell.fill = copy(source_cell.fill)
            if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
            if source_cell.number_format: target_cell.number_format = source_cell.number_format
# --------------------------------

def run_valet_automation():
    print(f"--- STARTING AUTOMATION: {datetime.datetime.now()} ---")
    downloaded_files = []
    
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) 
        root_folder = inbox.Parent
        
        target_folder = root_folder.Folders.Item("Revenue Tracking")
        print(f"TROUBLESHOOT: Accessed folder: {target_folder.Name}")

        now = datetime.datetime.now()
        first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        date_filter = first_of_month.strftime("%m/%d/%Y %I:%M %p")
        print(f"TROUBLESHOOT: Searching for emails received on or after {date_filter}")
        
        sender_email = "noreply@notifications.parkingmgt.com"
        sFilter = f"[SenderEmailAddress] = '{sender_email}' AND [ReceivedTime] >= '{date_filter}'"
        
        messages = target_folder.Items.Restrict(sFilter)
        messages.Sort("[ReceivedTime]", True)
        
        print(f"TROUBLESHOOT: Found {messages.Count} potential emails in Revenue Tracking.")

        found_recent_step2 = False
        found_recent_step3 = False

        for msg in messages:
            try:
                subject = msg.Subject
                subject_lower = subject.lower()
                
                if "trilogy" in subject_lower and "huntsville" in subject_lower:
                    if "step#1" in subject_lower or "step 1" in subject_lower:
                        prefix = "Step1_"
                    elif "step#2" in subject_lower or "step 2" in subject_lower:
                        if found_recent_step2:
                            continue  
                        prefix = "Step2_"
                        found_recent_step2 = True
                    elif "validation summary report" in subject_lower:
                        if found_recent_step3:
                            continue
                        prefix = "Step3_"
                        found_recent_step3 = True
                    else:
                        continue 

                    received_date = msg.ReceivedTime.strftime("%Y-%m-%d")
                    has_attachment = False
                    
                    for attachment in msg.Attachments:
                        if attachment.FileName.lower().endswith((".pdf", ".csv")):
                            has_attachment = True
                            unique_filename = f"{prefix}{received_date}_{attachment.FileName}"
                            pdf_path = os.path.join(current_folder, unique_filename)
                            
                            if not os.path.exists(pdf_path):
                                attachment.SaveAsFile(pdf_path)
                                downloaded_files.append(pdf_path)
                                print(f"SUCCESS: Downloaded '{unique_filename}'")
                            else:
                                downloaded_files.append(pdf_path)
                            break 
                            
            except Exception as e:
                pass 
        
        # --- D184 DOWNLOAD LOGIC ---
        print(f"TROUBLESHOOT: Searching Inbox for D184 emails...")
        inbox_messages = inbox.Items
        inbox_messages.Sort("[ReceivedTime]", True)
        
        d184_pdf_path = None
        for msg in inbox_messages:
            try:
                subject_lower = msg.Subject.lower() if msg.Subject else ""
                
                for attachment in msg.Attachments:
                    filename_lower = attachment.FileName.lower()
                    if filename_lower.endswith(".pdf") and ("d184" in subject_lower or "d184" in filename_lower):
                        current_date_str = datetime.datetime.now().strftime("%Y-%m-%d")
                        unique_filename = f"D184-{current_date_str}.pdf"
                        full_path = os.path.join(current_folder, unique_filename)
                        
                        if not os.path.exists(full_path):
                            attachment.SaveAsFile(full_path)
                            print(f"SUCCESS: Downloaded '{unique_filename}' from Inbox")
                        else:
                            print(f"TROUBLESHOOT: '{unique_filename}' already exists. Using existing file.")
                        
                        d184_pdf_path = full_path
                        break 
                
                if d184_pdf_path:
                    downloaded_files.append(d184_pdf_path)
                    break 
            except Exception as e:
                pass

    except Exception as e:
        print(f"TROUBLESHOOT ERROR (Outlook): {e}")

    if not downloaded_files:
        print("RESULT: No new data found in Outlook. Exiting.")
        return

    excel_path = r"C:\Users\AltonJones\Desktop\Automation test\2026 RSS-102742 Trilogy Hotel Huntsville.xlsx"
    
    try:
        print(f"TROUBLESHOOT: Opening workbook at {excel_path}")
        wb_save = openpyxl.load_workbook(excel_path)

        for filepath in downloaded_files:
            filename = os.path.basename(filepath)
            print(f"\n--- Processing: {filename} ---")
            
            # --- STEP #1 LOGIC ---
            if filename.startswith("Step1_") and filename.endswith(".pdf"):
                with pdfplumber.open(filepath) as pdf:
                    text = pdf.pages[0].extract_text()
                    
                date_match = re.search(r'Overnight Revenue Summary\s+([A-Z][a-z]{2})\s+(\d{1,2})', text, re.IGNORECASE)
                if not date_match: 
                    continue

                month_str = date_match.group(1).capitalize()
                day_num = int(date_match.group(2))
                
                vehicles = re.findall(r'Vehicle Count\s+(\d+)', text)
                revenues = re.findall(r'Revenue\s+\$?([\d,]+\.\d{2})', text)
                
                if len(vehicles) >= 2 and len(revenues) >= 2:
                    dv = int(vehicles[0])
                    dr = float(revenues[0].replace(',', ''))
                    ov = int(vehicles[1])
                    or_rev = float(revenues[1].replace(',', ''))

                    sheet_map = {"Jan": "RSS_Jan", "Feb": "RSS_Feb", "Mar": "RSS_Mar", "Apr": "RSS_Apr", "May": "RSS_May", "Jun": "RSS_June", "Jul": "RSS_July", "Aug": "RSS_Aug", "Sep": "RSS_Sept", "Oct": "RSS_Oct", "Nov": "RSS_Nov", "Dec": "RSS_Dec"}
                    sheet_name = sheet_map.get(month_str, f"RSS_{month_str}")
                    
                    if sheet_name in wb_save.sheetnames:
                        ws = wb_save[sheet_name]
                        
                        dv_row, dr_row = 73, 74
                        ov_row, or_rev_row = 79, 80
                        
                        for row in range(1, 150):
                            val_c = str(ws.cell(row=row, column=3).value).strip()
                            if val_c == "Valet Daily Revenue Collected by PMC":
                                dv_row, dr_row = row, row + 1
                            elif val_c == "Valet Overnight Revenue Collected by PMC":
                                ov_row, or_rev_row = row, row + 1

                        target_col = day_num + 4
                        
                        if not is_cell_merged(ws, dv_row, target_col): ws.cell(row=dv_row, column=target_col).value = dv
                        if not is_cell_merged(ws, dr_row, target_col): ws.cell(row=dr_row, column=target_col).value = dr
                        if not is_cell_merged(ws, ov_row, target_col): ws.cell(row=ov_row, column=target_col).value = ov
                        if not is_cell_merged(ws, or_rev_row, target_col): ws.cell(row=or_rev_row, column=target_col).value = or_rev
                            
                        print(f"SUCCESS: Daily & Overnight Data for {month_str} {day_num} updated in Column {target_col}")

            # --- D184 LOGIC ---
            elif filename.startswith("D184-") and filename.endswith(".pdf"):
                sheet_map_by_num = {1: "RSS_Jan", 2: "RSS_Feb", 3: "RSS_Mar", 4: "RSS_Apr", 5: "RSS_May", 6: "RSS_June", 7: "RSS_July", 8: "RSS_Aug", 9: "RSS_Sept", 10: "RSS_Oct", 11: "RSS_Nov", 12: "RSS_Dec"}
                daily_totals = {}
                
                with pdfplumber.open(filepath) as pdf:
                    for page in pdf.pages:
                        table = page.extract_table()
                        if table:
                            for row in table:
                                if row and len(row) >= 5 and row[0]:
                                    date_str = str(row[0]).strip().replace('\n', '')
                                    match = re.match(r'^(\d{2})-(\d{2})-(\d{2})$', date_str)
                                    if match:
                                        month_num = int(match.group(1))
                                        day_num = int(match.group(2))
                                        amount_str = str(row[-1]).strip().replace('$', '').replace(',', '')
                                        try:
                                            amount = float(amount_str)
                                            key = (month_num, day_num)
                                            daily_totals[key] = daily_totals.get(key, 0.0) + amount
                                        except ValueError:
                                            pass
                        else:
                            # Fallback if table extraction fails
                            text = page.extract_text()
                            if text:
                                for line in text.split('\n'):
                                    match = re.search(r'^(\d{2})-(\d{2})-(\d{2})\s+.*\s+([\d,]+\.\d{2})$', line.strip())
                                    if match:
                                        month_num = int(match.group(1))
                                        day_num = int(match.group(2))
                                        amount = float(match.group(4).replace(',', ''))
                                        key = (month_num, day_num)
                                        daily_totals[key] = daily_totals.get(key, 0.0) + amount

                for (month_num, day_num), total_amount in daily_totals.items():
                    sheet_name = sheet_map_by_num.get(month_num)
                    if sheet_name and sheet_name in wb_save.sheetnames:
                        ws = wb_save[sheet_name]
                        target_col = day_num + 4
                        target_row = 36
                        
                        if not is_cell_merged(ws, target_row, target_col):
                            ws.cell(row=target_row, column=target_col).value = total_amount
                            print(f"SUCCESS: Wrote D184 total ${total_amount:.2f} for {sheet_name} Day {day_num} to Row {target_row}, Column {target_col}")

            # --- STEP #2 LOGIC ---
            elif filename.startswith("Step2_"):
                sheet_name = "Session Discrepancy"
                if sheet_name in wb_save.sheetnames:
                    ws_disc = wb_save[sheet_name]
                    
                    existing_tickets = set()
                    for r in range(10, ws_disc.max_row + 1):
                        val = ws_disc.cell(row=r, column=3).value
                        if val is not None:
                            if isinstance(val, float): existing_tickets.add(str(int(val)))
                            else: existing_tickets.add(str(val).strip())

                    last_data_row = get_true_bottom_row(ws_disc, 10)
                    last_month = None
                    if last_data_row >= 10:
                        last_arrival_val = ws_disc.cell(row=last_data_row, column=5).value
                        last_month = get_month(last_arrival_val)

                    next_empty_row = get_next_safe_row(ws_disc, last_data_row + 1)
                    extracted_data = []
                    
                    if filename.endswith(".pdf"):
                        with pdfplumber.open(filepath) as pdf:
                            for page in pdf.pages:
                                table = page.extract_table()
                                if table: extracted_data.extend(table)
                    else:
                        with open(filepath, newline='', encoding='utf-8') as f:
                            reader = csv.reader(f)
                            extracted_data = list(reader)

                    headers = []
                    processing_overnights = False
                    
                    for row_data in extracted_data:
                        row_data = [str(x).strip().replace('\n', ' ') if x else "" for x in row_data]
                        
                        if all(x == "" for x in row_data): continue
                            
                        row_text = " ".join(row_data)
                        if "Overnight Valet" in row_text:
                            processing_overnights = True
                            continue
                        elif "Daily Valet" in row_text:
                            break
                            
                        if not processing_overnights: continue
                        if "Ticket" in row_data and "Guest Name" in row_data:
                            headers = row_data
                            continue
                        if "Total" in row_data: continue
                            
                        if headers and len(row_data) > 0 and row_data[0].isdigit():
                            ticket = row_data[headers.index("Ticket")] if "Ticket" in headers else ""
                            
                            ticket_str = str(ticket).strip()
                            if ticket_str in existing_tickets: continue
                            
                            guest_name = row_data[headers.index("Guest Name")] if "Guest Name" in headers else ""
                            arrival = row_data[headers.index("Arrival")] if "Arrival" in headers else ""
                            
                            depart = ""
                            if "Departure" in headers: depart = row_data[headers.index("Departure")]
                            elif "Depart" in headers: depart = row_data[headers.index("Depart")]
                            
                            room = ""
                            if "Room Number" in headers and headers.index("Room Number") < len(row_data): 
                                room = row_data[headers.index("Room Number")]
                                
                            try:
                                room_val = int(room)
                            except ValueError:
                                try:
                                    room_val = float(room)
                                except ValueError:
                                    room_val = room
                            
                            collected = "0"
                            if "Collected" in headers and headers.index("Collected") < len(row_data): 
                                collected = row_data[headers.index("Collected")].replace("$","").replace(",","")
                            
                            uncollected = "0"
                            if "Uncollected" in headers and headers.index("Uncollected") < len(row_data): 
                                uncollected = row_data[headers.index("Uncollected")].replace("$","").replace(",","")
                            
                            current_month = get_month(arrival)
                            if current_month and last_month and current_month != last_month:
                                next_empty_row = get_next_safe_row(ws_disc, next_empty_row + 1)
                            
                            if current_month: last_month = current_month

                            ws_disc.cell(row=next_empty_row, column=3).value = int(ticket) if ticket.isdigit() else ticket
                            ws_disc.cell(row=next_empty_row, column=4).value = guest_name
                            ws_disc.cell(row=next_empty_row, column=5).value = arrival
                            ws_disc.cell(row=next_empty_row, column=6).value = depart
                            ws_disc.cell(row=next_empty_row, column=7).value = room_val
                            
                            try: ws_disc.cell(row=next_empty_row, column=8).value = float(collected)
                            except ValueError: pass
                                
                            try: ws_disc.cell(row=next_empty_row, column=9).value = float(uncollected)
                            except ValueError: pass

                            apply_row_formatting(ws_disc, 10, next_empty_row, 3, 9)
                            
                            print(f"SUCCESS: Wrote ticket {ticket} to safe row {next_empty_row}")
                            
                            existing_tickets.add(ticket_str)
                            next_empty_row = get_next_safe_row(ws_disc, next_empty_row + 1)

            # --- STEP #3 LOGIC (BTR Tracker) ---
            elif filename.startswith("Step3_") and filename.endswith(".csv"):
                sheet_name = "BTR Tracker"
                if sheet_name in wb_save.sheetnames:
                    ws_btr = wb_save[sheet_name]
                    
                    existing_names = set()
                    for r in range(10, ws_btr.max_row + 1):
                        val = ws_btr.cell(row=r, column=4).value
                        if val is not None:
                            existing_names.add(str(val).strip().lower())

                    last_data_row = get_true_bottom_row(ws_btr, 10)
                    last_month = None
                    if last_data_row >= 10:
                        last_arrival_val = ws_btr.cell(row=last_data_row, column=6).value
                        last_month = get_month(last_arrival_val)

                    next_empty_row = get_next_safe_row(ws_btr, last_data_row + 1)

                    extracted_data = []
                    with open(filepath, newline='', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            extracted_data.append(row)

                    valid_rows = []
                    for row in extracted_data:
                        if len(row) >= 9 and row[1].strip().isdigit():
                            validation_str = row[6].strip().replace('\n', ' ')
                            if "bill to room 2026 (btr26)" in validation_str.lower():
                                valid_rows.append(row)

                    valid_rows.sort(key=lambda x: (
                        x[6].strip(), 
                        parse_csv_date(x[4]), 
                        parse_csv_date(x[5]), 
                        int(x[1].strip())
                    ))

                    for row_data in valid_rows:
                        guest_name = row_data[2].strip()
                        name_key = guest_name.lower()
                        
                        if name_key in existing_names:
                            print(f"TROUBLESHOOT: Skipping duplicate name '{guest_name}'")
                            continue

                        ticket_str = row_data[1].strip()
                        room = row_data[3].strip()
                        
                        try:
                            room_val = int(room)
                        except ValueError:
                            try:
                                room_val = float(room)
                            except ValueError:
                                room_val = room
                                
                        arrival_raw = row_data[4].strip()
                        depart_raw = row_data[5].strip()
                        validation = row_data[6].strip().replace('\n', ' ')
                        amount = row_data[8].strip().replace('$', '').replace(',', '')
                        
                        arrival_dt = parse_date_to_obj(arrival_raw)
                        depart_dt = parse_date_to_obj(depart_raw)

                        current_month = get_month(arrival_raw)
                        if current_month and last_month and current_month != last_month:
                            print(f"TROUBLESHOOT: Month boundary detected in BTR ({last_month} -> {current_month}). Adding blank row.")
                            next_empty_row = get_next_safe_row(ws_btr, next_empty_row + 1)
                        
                        if current_month:
                            last_month = current_month

                        ws_btr.cell(row=next_empty_row, column=3).value = int(ticket_str)         
                        ws_btr.cell(row=next_empty_row, column=4).value = guest_name              
                        ws_btr.cell(row=next_empty_row, column=5).value = room_val                
                        ws_btr.cell(row=next_empty_row, column=6).value = arrival_dt if arrival_dt else arrival_raw 
                        ws_btr.cell(row=next_empty_row, column=7).value = depart_dt if depart_dt else depart_raw   
                        ws_btr.cell(row=next_empty_row, column=8).value = validation              
                        ws_btr.cell(row=next_empty_row, column=9).value = "Yes"                   
                        
                        try:
                            ws_btr.cell(row=next_empty_row, column=10).value = float(amount)      
                        except ValueError:
                            ws_btr.cell(row=next_empty_row, column=10).value = amount
                        
                        apply_row_formatting(ws_btr, 10, next_empty_row, 3, 10)

                        print(f"SUCCESS: Wrote BTR ticket for {guest_name} to safe row {next_empty_row}")
                        
                        existing_names.add(name_key) 
                        next_empty_row = get_next_safe_row(ws_btr, next_empty_row + 1)

        wb_save.save(excel_path)
        wb_save.close()
        
        print("\nSUCCESS: Workbook saved successfully. Cleaning up temp files...")

        for filepath in downloaded_files:
            try: 
                os.remove(filepath)
            except: pass
            
        print("Done.")

    except Exception as e:
        import traceback
        print(f"CRITICAL ERROR: {e}")
        traceback.print_exc() 

if __name__ == "__main__":
    run_valet_automation()
